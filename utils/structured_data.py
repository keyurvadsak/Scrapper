"""
Structured data utilities for exporting scraped data to Excel format.
Creates beautifully formatted Excel workbooks with color-coded tables.
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.logger_config import setup_logger

logger = setup_logger(__name__)

# Color scheme for different tag types (matching your image)
TAG_COLORS = {
    'page_info': 'D3D3D3',      # Light gray
    'h1': '4472C4',              # Blue (Headings)
    'h2': 'C55A5A',              # Red (Sub Heading)
    'h3': 'C55A5A',              # Red (Sub Heading)
    'h4': 'C55A5A',              # Red (Sub Heading)
    'h5': 'C55A5A',              # Red (Sub Heading)
    'h6': 'C55A5A',              # Red (Sub Heading)
    'p': '9DC183',               # Green (Paragraphs)
    'li': '4472C4',              # Blue (List Item)
    'th': '9B7FB5',              # Purple (table heading)
    'td': '70C0CF',              # Cyan (table data)
    'a': 'F4B183',               # Orange (Internal Links)
    'external_links': '4472C4',  # Blue (External Links)
    'img': 'C55A5A',             # Red (Image)
    'video': '9DC183',           # Green (Video)
    'audio': '9B7FB5',           # Purple (Audio)
    'overview': '38cb82',          # Light green (Overview)
}

TAG_LABELS = {
    'h1': 'Headings',
    'h2': 'Sub Heading',
    'h3': 'Sub Heading',
    'h4': 'Sub Heading',
    'h5': 'Sub Heading',
    'h6': 'Sub Heading',
    'p': 'Paragraphs',
    'li': 'List Item',
    'th': 'table heading',
    'td': 'table data',
    'a': 'Internal Links',
    'img': 'Image',
    'video': 'Video',
    'audio': 'Audio',
}


def apply_header_style(cell, color_hex):
    """Apply header styling to a cell."""
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def apply_data_style(cell):
    """Apply data cell styling."""
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def create_overview_sheet(ws, website_name, website_url, pages_data):
    """
    Create the overview sheet with website info and page index.
    
    Args:
        ws: Worksheet object
        website_name: Name of the website
        website_url: Base URL of the website
        pages_data: List of dicts with 'name' and 'url' keys
    """
    ws.title = "Overview"
    
    # Website info section
    ws['A1'] = 'Website Name'
    ws['B1'] = website_name
    ws['A2'] = 'Website URL'
    ws['B2'] = website_url
    
    apply_header_style(ws['A1'], TAG_COLORS['page_info'])
    apply_header_style(ws['A2'], TAG_COLORS['page_info'])
    apply_data_style(ws['B1'])
    apply_data_style(ws['B2'])

    ws['A4'] = 'Click on page name to view details'
    apply_header_style(ws['A4'], TAG_COLORS['overview'])
    
    # Page index table
    ws['A6'] = 'Page Name'
    ws['B6'] = 'Page URL'
    
    apply_header_style(ws['A6'], TAG_COLORS['th'])
    apply_header_style(ws['B6'], TAG_COLORS['th'])
    
    # Add page data
    for idx, page in enumerate(pages_data, start=7):
        page_number = idx - 4
        sheet_name = f"Page_{page_number}"
        # ws[f'A{idx}'] = f"page_{idx-4}"
        cell = ws[f'A{idx}']
        cell.value = f"page_{page_number}"
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.style = "Hyperlink"
        ws[f'B{idx}'] = page['url']
        apply_data_style(ws[f'A{idx}'])
        apply_data_style(ws[f'B{idx}'])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60
    
    logger.info(f"Created overview sheet with {len(pages_data)} pages")


def create_page_sheet(ws, page_name, page_url, tags_data):
    """
    Create a sheet for a single page with color-coded tag tables.
    
    Args:
        ws: Worksheet object
        page_name: Name of the page
        page_url: URL of the page
        tags_data: Dictionary of tag data from database
    """
    # current_row = 1

    
    ws['A1'] = "← Back to Overview"
    ws['A1'].hyperlink = "#'Overview'!A1"
    ws['A1'].style = "Hyperlink"

    ws.merge_cells('A1:A1')
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    apply_header_style(ws['A1'], TAG_COLORS['overview'])

    
    # Page info section
    ws['A3'] = 'Page Name'
    ws['B3'] = page_name
    ws['A4'] = 'Page URL'
    ws['B4'] = page_url
    # current_row = 5
    
    apply_header_style(ws['A3'], TAG_COLORS['page_info'])
    apply_header_style(ws['A4'], TAG_COLORS['page_info'])
    apply_data_style(ws['B3'])
    apply_data_style(ws['B4'])

    
    current_row = 6
    
    # Process each tag type
    tag_order = ['h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'p', 'li', 'th', 'td', 'a', 'img', 'video', 'audio']
    
    for tag in tag_order:
        if tag not in tags_data or not tags_data[tag]:
            continue
        
        tag_label = TAG_LABELS.get(tag, tag.upper())
        tag_color = TAG_COLORS.get(tag, '4472C4')
        tag_elements = tags_data[tag]
        
        # Skip if no data
        if not tag_elements:
            current_row += 1
            continue
        
        # Determine columns based on tag type
        if tag in ['a']:  # Links have text and href
            # Header row
            ws[f'A{current_row}'] = tag_label
            ws[f'B{current_row}'] = 'Text'
            ws[f'C{current_row}'] = 'href'
            
            apply_header_style(ws[f'A{current_row}'], tag_color)
            apply_header_style(ws[f'B{current_row}'], tag_color)
            apply_header_style(ws[f'C{current_row}'], tag_color)
            
            current_row += 1
            
            # Data rows
            for key, value in tag_elements.items():
                ws[f'A{current_row}'] = tag.upper()
                ws[f'B{current_row}'] = value.get('text', '') if isinstance(value, dict) else ''
                ws[f'C{current_row}'] = value.get('href', '') if isinstance(value, dict) else ''
                
                apply_data_style(ws[f'A{current_row}'])
                apply_data_style(ws[f'B{current_row}'])
                apply_data_style(ws[f'C{current_row}'])
                current_row += 1
                
        elif tag in ['img', 'video', 'audio']:  # Media have src and alt
            # Header row
            ws[f'A{current_row}'] = tag_label
            ws[f'B{current_row}'] = 'src'
            ws[f'C{current_row}'] = 'alt'
            
            apply_header_style(ws[f'A{current_row}'], tag_color)
            apply_header_style(ws[f'B{current_row}'], tag_color)
            apply_header_style(ws[f'C{current_row}'], tag_color)
            
            current_row += 1
            
            # Data rows
            for key, value in tag_elements.items():
                ws[f'A{current_row}'] = tag.upper()
                ws[f'B{current_row}'] = value.get('src', '') if isinstance(value, dict) else ''
                ws[f'C{current_row}'] = value.get('alt', '') if isinstance(value, dict) else ''
                
                apply_data_style(ws[f'A{current_row}'])
                apply_data_style(ws[f'B{current_row}'])
                apply_data_style(ws[f'C{current_row}'])
                current_row += 1
                
        else:  # Text-only tags
            # Header row
            ws[f'A{current_row}'] = tag_label
            ws[f'B{current_row}'] = 'Text'
            
            apply_header_style(ws[f'A{current_row}'], tag_color)
            apply_header_style(ws[f'B{current_row}'], tag_color)
            
            current_row += 1
            
            # Data rows
            for key, value in tag_elements.items():
                ws[f'A{current_row}'] = tag.upper()
                ws[f'B{current_row}'] = value if isinstance(value, str) else ''
                
                apply_data_style(ws[f'A{current_row}'])
                apply_data_style(ws[f'B{current_row}'])
                current_row += 1
        
        # Add spacing between sections
        current_row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 50
    
    logger.info(f"Created sheet for page: {page_name}")


def sanitize_sheet_name(name, max_length=31):
    """
    Sanitize sheet name to comply with Excel requirements.
    - Max 31 characters
    - No special characters: : \\ / ? * [ ]
    """
    # Remove invalid characters
    invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
    for char in invalid_chars:
        name = name.replace(char, '_')
    
    # Truncate to max length
    if len(name) > max_length:
        name = name[:max_length]
    
    # Ensure not empty
    if not name:
        name = "Sheet"
    
    return name


def create_excel_workbook(website_data, pages_data, output_path):
    """
    Create an Excel workbook for a website.
    
    Args:
        website_data: Dict with 'name' and 'base_url'
        pages_data: List of dicts with 'name', 'url', and 'tags' (JSONB data)
        output_path: Path where to save the Excel file
    """
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create overview sheet
    overview_ws = wb.create_sheet("Overview", 0)
    pages_info = [{'name': f"Page_{idx}", 'url': p['url']} for idx,p in enumerate(pages_data, start=1)]
    create_overview_sheet(overview_ws, website_data['name'], website_data['base_url'], pages_info)
    
    # Create a sheet for each page
    for idx, page in enumerate(pages_data, start=1):
        # Sanitize sheet name   
        sheet_name = sanitize_sheet_name(f"Page_{idx}")
        
        # Ensure unique sheet names
        # if sheet_name in wb.sheetnames:
        #     sheet_name = sanitize_sheet_name(f"{page['name']}_{idx}")
        
        page_ws = wb.create_sheet(sheet_name)
        create_page_sheet(page_ws, page['name'], page['url'], page['tags'])
    
    # Save workbook
    wb.save(output_path)
    logger.info(f"Excel workbook saved to: {output_path}")
    print(f"✅ Excel file created: {output_path}")
    
    return output_path